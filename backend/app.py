import eventlet
eventlet.monkey_patch()

from flask import Flask, request, jsonify
from flask_socketio import SocketIO
from flask_cors import CORS, cross_origin # <-- Add cross_origin here
import serial
from serial.tools import list_ports # <-- ADD THIS LINE
import threading
import time
import base64
import sqlite3
import logging
import os
from openpyxl import Workbook, load_workbook
import json
import time # For simulating real-time sensor data
#import random # For simulating sensor data
import serial # For serial communication
import threading # For running serial reading in a separate thread
import time # For time.sleep
from fpdf import FPDF



app = Flask(__name__)

# IMPORTANT: Removed Flask session-related secret_key for local storage auth.
CORS(app, resources={r"/*": {"origins": "*"}}, supports_credentials=True) # Enable CORS for all routes (supports_credentials no longer needed with localStorage auth)
socketio = SocketIO(app, cors_allowed_origins="*", async_mode='eventlet')
@socketio.on('connect')

def handle_connect():
    print("🟢 A client connected via Socket.IO")
    print(f"Client connected: {request.sid}")

# cors_allowed_origins="*" allows connections from any origin.
# For production, you should restrict this to your frontend's specific origin (e.g., "http://localhost:3000").
# async_mode='eventlet' tells SocketIO to use eventlet for asynchronous operations, which you installed.
logging.basicConfig(level=logging.DEBUG)

# Global dictionary to simulate sensor data for ongoing tests
current_test_sensor_data = {}
# Global variables for serial communication and live test management
ser = None  # Will hold the pyserial Serial object
sensor_reading_thread = None  # Will hold the Thread object for reading serial data
is_test_active = False  # Flag to control when to process and emit sensor data
current_test_id_for_live_data = None  # Stores the test_id for currently active walk test

# --- Serial Port Reading Thread Function ---
# --- Helper Function for Serial Port Auto-Detection ---
# --- Helper Function for Serial Port Auto-Detection ---
def auto_detect_esp_port():
    """
    Attempts to auto-detect the serial port of an ESP32.
    It looks for common Vendor IDs (VIDs) and Product IDs (PIDs) associated with ESP32s
    (e.g., CP210x, CH340/CH341 USB-to-serial converters).
    """
    # Common VIDs and PIDs for ESP32s / their USB-to-Serial chips
    # CP210x: Silicon Labs (e.g., ESP32-DevKitC, some generic ESP32 boards)
    CP210X_VID = 0x10C4 # Silicon Labs
    CP210X_PID = 0xEA60 # CP2102/CP2104

    # CH340/CH341: QinHeng Electronics (common on cheaper ESP32 boards)
    CH340_VID = 0x1A86 # QinHeng Electronics
    CH340_PID = 0x7523 # CH340G / CH341A

    # NEW: ESP32-C3 specific VID/PID (identified from your debug output: 303A:1001)
    ESP32_C3_VID = 0x303A
    ESP32_C3_PID = 0x1001

    found_port = None
    print("INFO: Attempting to auto-detect ESP32 serial port...")

    # Iterate over all available serial ports
    for port in list_ports.comports():
        # Keep this debug print for now, it's very useful!
        print(f"DEBUG: Found port: {port.device}, Desc: {port.description}, HWID: {port.hwid}, VID: {port.vid}, PID: {port.pid}")

        # Check by Vendor ID and Product ID
        if (port.vid == CP210X_VID and port.pid == CP210X_PID) or \
           (port.vid == CH340_VID and port.pid == CH340_PID) or \
           (port.vid == ESP32_C3_VID and port.pid == ESP32_C3_PID): # <-- ADDED ESP32-C3 PID/VID CHECK
            print(f"INFO: Auto-detected ESP32 port (VID/PID match): {port.device}")
            found_port = port.device
            break # Found one, take the first one

        # Fallback: Check by common descriptions (less reliable than VID/PID)
        # This is less critical now that we have the exact VID/PID for C3, but good to keep.
        if "CP210" in port.description or "USB-SERIAL CH340" in port.description or "ESP32-C3" in port.description:
            print(f"INFO: Auto-detected ESP32 port (Description match): {port.device}")
            found_port = port.device
            break # Found one, take the first one

    if found_port:
        return found_port
    else:
        print("WARNING: No ESP32 serial port auto-detected based on common VIDs/PIDs or descriptions.")
        print("Please ensure your Master ESP32 is connected and its drivers are installed.")
        print("If detection fails, you might need to manually set the COM_PORT variable.")
        return None

# --- Serial Port Reading Thread Function ---
def read_from_serial_port():
    """
    Reads data continuously from the serial port, parses it, maps it to UI groups,
    emits it via WebSocket, and optionally stores it to the database.
    This function is intended to run in a separate thread.
    """
    global ser, is_test_active, current_test_id_for_live_data, socketio, app

    BAUD_RATE = 115200 # Matches your ESP code

    # This block now attempts to find and open the serial port initially.
    # The while loop will then continuously try if it fails or disconnects.
    COM_PORT_TO_ATTEMPT = auto_detect_esp_port() # Get the auto-detected port

    try:
        if COM_PORT_TO_ATTEMPT: # If a port was auto-detected
            ser = serial.Serial(COM_PORT_TO_ATTEMPT, BAUD_RATE, timeout=1)
            print(f"INFO: Serial port {COM_PORT_TO_ATTEMPT} opened successfully.")
            # Give the ESP a moment to reset after port opening
            time.sleep(2)
            # Send 'nope' initially to ensure ESP doesn't send data until explicitly 'start'
            ser.write(b'nope\n')
            print("INFO: Sent 'nope' command to ESP on startup.")
        else:
            # If auto-detection failed initially, ser remains None.
            # The main while loop will handle continuous retries for detection and opening.
            ser = None
            print("WARNING: Auto-detection failed during initial startup. The serial thread will continuously try to find and open the ESP32 port.")

    except serial.SerialException as e:
        print(f"ERROR: Could not open initial serial port {COM_PORT_TO_ATTEMPT or ' (no port detected)'}: {e}")
        ser = None # Ensure ser is None if opening fails
    except Exception as e:
        print(f"CRITICAL ERROR during initial serial setup: {e}")
        ser = None

    # Main continuous reading loop
    while True:
        if ser and ser.is_open:
            try:
                # Read a line from the serial port
                line = ser.readline().decode('utf-8').strip()

                if line:
                    # Handle known plain text messages from ESP startup/status
                    #if "Access Point started" in line or "IP Address:" in line or "Connected to WiFi" in line:
                    #   print(f"INFO: Received ESP32 status message: {line}")
                    #   continue # Skip to next line read, no further processing needed for this line

                    # Assume remaining lines are comma-separated sensor values
                    values_str = line.split(',')

                    # Ensure we have 12 values (6 from Left, 6 from Right)
                    if len(values_str) == 12:
                        try:
                            # Convert string values to floats for calculations
                            raw_sensor_values = [float(v) for v in values_str]

                            # --- Hypothetical Mapping to UI Groups (YOU MUST CONFIRM/ADJUST THIS) ---
                            # This mapping is based on the assumption that the first 6 values are Left shoe
                            # sensors (L1-L6) and the next 6 are Right shoe sensors (R1-R6).
                            # Adjust the indices and aggregation method (e.g., sum, average, direct)
                            # based on how your physical sensors map to G1, G2, G3.

                            # Extract Left and Right Shoe raw values
                            L1, L2, L3, L4, L5, L6 = raw_sensor_values[0:6]
                            R1, R2, R3, R4, R5, R6 = raw_sensor_values[6:12]

                            # Example mapping (adjust as per your sensor layout and logic):
                            left_g1 = (L1 + L2) / 2.0 if (L1 + L2) > 0 else 0.0
                            left_g2 = (L3 + L4) / 2.0 if (L3 + L4) > 0 else 0.0
                            left_g3 = (L5 + L6) / 2.0 if (L5 + L6) > 0 else 0.0

                            right_g1 = (R1 + R2) / 2.0 if (R1 + R2) > 0 else 0.0
                            right_g2 = (R3 + R4) / 2.0 if (R3 + R4) > 0 else 0.0
                            right_g3 = (R5 + R6) / 2.0 if (R5 + R6) > 0 else 0.0

                            processed_data = {
                                'left_g1': round(left_g1, 2),
                                'left_g2': round(left_g2, 2),
                                'left_g3': round(left_g3, 2),
                                'right_g1': round(right_g1, 2),
                                'right_g2': round(right_g2, 2),
                                'right_g3': round(right_g3, 2)
                            }

                            # --- Conditional Data Emission & Storage ---
                            # Only emit and store if a test is currently active
                            if is_test_active and current_test_id_for_live_data is not None:
                                # Emit to Frontend via WebSocket
                                with app.app_context():
                                    socketio.emit('sensor_data', raw_sensor_values)
                                if current_test_id_for_live_data in current_test_sensor_data:
                                    current_test_sensor_data[current_test_id_for_live_data]['data'].append(raw_sensor_values)
                                # Store to Database (single row for all 12 raw and 6 processed sensor values)
                                '''try:
                                    conn = sqlite3.connect('spinalysis.db')
                                    c = conn.cursor()
                                    # Insert all 12 raw sensor values AND 6 processed grouped values into a single row
                                    c.execute(
                                        INSERT INTO sensor_data (
                                            test_id, timestamp,
                                            L1, L2, L3, L4, L5, L6,
                                            R1, R2, R3, R4, R5, R6,
                                            left_g1, left_g2, left_g3,
                                            right_g1, right_g2, right_g3
                                        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                                    , (
                                        current_test_id_for_live_data,
                                        int(time.time()), # Current timestamp
                                        raw_sensor_values[0], raw_sensor_values[1], raw_sensor_values[2],
                                        raw_sensor_values[3], raw_sensor_values[4], raw_sensor_values[5],
                                        raw_sensor_values[6], raw_sensor_values[7], raw_sensor_values[8],
                                        raw_sensor_values[9], raw_sensor_values[10], raw_sensor_values[11],
                                        processed_data['left_g1'], processed_data['left_g2'], processed_data['left_g3'],
                                        processed_data['right_g1'], processed_data['right_g2'], processed_data['right_g3']
                                    ))
                                    conn.commit()
                                    conn.close()
                                    # print("DEBUG: All 18 sensor values stored to DB successfully.") # Optional: add debug print
                                except sqlite3.Error as db_e:
                                    print(f"ERROR: Database insertion failed: {db_e}")
                                except Exception as general_e:
                                    print(f"ERROR: General error during database insertion: {general_e}")'''
                        except ValueError as ve:
                            # This will catch errors if values are not numbers (e.g., from malformed data)
                            print(f"WARNING: Could not convert data to float (non-numeric value found): {ve} in line: {line}")
                        except IndexError as ie:
                            # This should ideally not be hit if len(values_str) == 12 check passes,
                            # but kept for robustness for unexpected line formats.
                            print(f"WARNING: Insufficient values after split: {ie} in line: {line}")
                    else:
                        # This catches any other lines that are not 12 comma-separated values
                        print(f"WARNING: Received line with unexpected format or incorrect number of values ({len(values_str)}): {line}")
            except serial.SerialTimeoutException:
                pass
            except serial.SerialException as e:
                print(f"ERROR: Serial port error during read: {e}. Attempting to re-open.")
                if ser: ser.close() # Close if it was open
                ser = None
                time.sleep(5) # Wait before trying to open again
            except Exception as e:
                print(f"CRITICAL ERROR in serial thread: {e}")
        else:
            # If serial port is not open, wait and try to open it again
            print(f"DEBUG: Serial port not open. Attempting to auto-detect and open...")
            COM_PORT_TO_ATTEMPT = auto_detect_esp_port()
            if COM_PORT_TO_ATTEMPT:
                try:
                    ser = serial.Serial(COM_PORT_TO_ATTEMPT, BAUD_RATE, timeout=1)
                    print(f"INFO: Serial port {COM_PORT_TO_ATTEMPT} opened successfully after retry.")
                    time.sleep(2)
                    ser.write(b'nope\n')
                except serial.SerialException as e:
                    print(f"WARNING: Still cannot open serial port {COM_PORT_TO_ATTEMPT}: {e}")
                    ser = None
            else:
                # Still no port detected, wait before next auto-detection attempt
                time.sleep(5)

        time.sleep(0.01) # Small delay to prevent busy-waiting

def init_db():
    """Initializes the SQLite database and tables if they don't exist."""
    conn = sqlite3.connect('spinalysis.db')
    c = conn.cursor()
    # Doctors table for authentication
    c.execute('''CREATE TABLE IF NOT EXISTS doctors (username TEXT PRIMARY KEY, password TEXT)''')
    # Patients table to store patient details, now linked to a doctor
    c.execute('''CREATE TABLE IF NOT EXISTS patients (
                    patient_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    first_name TEXT,
                    middle_name TEXT,
                    last_name TEXT,
                    age INTEGER,
                    hospital_number TEXT UNIQUE,
                    doctor_username TEXT,
                    FOREIGN KEY(doctor_username) REFERENCES doctors(username)
                )''')
    # Walk tests table to store details and collected sensor data for each test
    c.execute('''CREATE TABLE IF NOT EXISTS walk_tests (
                    test_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    patient_id INTEGER,
                    duration INTEGER,
                    walk_type TEXT,
                    start_time REAL,
                    end_time REAL,
                    sensor_data_json TEXT,
                    FOREIGN KEY(patient_id) REFERENCES patients(patient_id)
                )''')
    # Modified to include grouped pressure values (left_g1 to right_g3)
    # Changed L1-R6 to REAL type

    conn.commit()
    conn.close()

init_db() # Initialize the database when the application starts




@socketio.on('start_test_command')
def handle_start_test_command(data):
    """
    Handles the 'start_test_command' event from the frontend.
    It performs the following:
    1. Validates input
    2. Verifies patient belongs to doctor
    3. Inserts walk test into database
    4. Sets global variables for live data collection
    5. Sends 'start' to Master ESP
    6. Emits acknowledgment to frontend
    """
    global is_test_active, current_test_id_for_live_data, current_test_sensor_data, ser

    try:
        # Step 1: Extract and validate input fields
        patient_id = data.get('patient_id')
        duration = data.get('duration')  # seconds
        walk_type = data.get('walk_type')
        doctor_username = data.get('doctor_username')

        if not all([patient_id, duration, walk_type, doctor_username]):
            socketio.emit('test_started_ack', {
                'success': False,
                'message': 'Missing one or more required fields (patient_id, duration, walk_type, doctor_username)'
            })
            return

        # Step 2: Check if patient belongs to doctor
        conn = sqlite3.connect('spinalysis.db')
        c = conn.cursor()
        c.execute("SELECT 1 FROM patients WHERE patient_id = ? AND doctor_username = ?", (patient_id, doctor_username))
        if c.fetchone() is None:
            conn.close()
            socketio.emit('test_started_ack', {
                'success': False,
                'message': 'Unauthorized: Patient does not belong to this doctor'
            })
            return

        # Step 3: Insert walk test into DB
        start_time = int(time.time())  # UNIX timestamp
        empty_json = json.dumps([])  # Initialize sensor data as empty list
        c.execute("""
            INSERT INTO walk_tests (patient_id, duration, walk_type, start_time, sensor_data_json)
            VALUES (?, ?, ?, ?, ?)
        """, (patient_id, duration, walk_type, start_time, empty_json))
        test_id = c.lastrowid
        conn.commit()
        conn.close()

        # Step 4: Set global tracking variables
        current_test_id_for_live_data = test_id
        is_test_active = True
        current_test_sensor_data[test_id] = {
            'data': [],
            'start_time': start_time,
            'duration': duration,
            'walk_type': walk_type
        }

        # Step 5: Start ESP streaming
        if ser and ser.is_open:
            ser.write(b'start\n')
            print(f"INFO: Sent 'start' to Master ESP for test_id {test_id}")
        else:
            print("WARNING: Serial port not open. ESP not triggered.")
            # Optional: Reset global flags if critical
            is_test_active = False
            current_test_id_for_live_data = None

        # Step 6: Acknowledge to frontend
        socketio.emit('test_started_ack', {
            'success': True,
            'test_id': test_id,
            'message': 'Test started successfully.'
        })

    except Exception as e:
        print(f"CRITICAL ERROR in handle_start_test_command: {str(e)}")
        is_test_active = False
        current_test_id_for_live_data = None
        socketio.emit('test_started_ack', {
            'success': False,
            'message': f'Server error: {str(e)}'
        })



@socketio.on('stop_test_command')
def handle_stop_test_command(data):
    print(f"DEBUG: Entered handle_stop_test_command with data: {data}")

    global is_test_active, current_test_id_for_live_data, ser, app, current_test_sensor_data

    test_id = data.get('testId')
    doctor_username = data.get('doctor_username')

    if not test_id or not doctor_username:
        socketio.emit('test_stopped_ack', {
            'test_id': test_id,
            'success': False,
            'message': 'Missing testId or doctor_username'
        })
        return

    # Step 1: Stop sensor activity
    is_test_active = False
    current_test_id_for_live_data = None

    if ser and ser.is_open:
        ser.write(b'nope\n')
        print("INFO: Sent 'nope' command to Master ESP.")
    else:
        print("WARNING: Serial port not open. Cannot send 'nope' command to ESP.")

    conn = None
    try:
        conn = sqlite3.connect('spinalysis.db')
        c = conn.cursor()

        # Step 2: Authorization check
        c.execute("SELECT T.patient_id, T.walk_type FROM walk_tests T JOIN patients P ON T.patient_id = P.patient_id WHERE T.test_id = ? AND P.doctor_username = ?", (test_id, doctor_username))
        test_db_info = c.fetchone()

        if not test_db_info:
            socketio.emit('test_stopped_ack', {
                'test_id': test_id,
                'success': False,
                'message': 'Unauthorized or test not found'
            })
            return

        patient_id, walk_type = test_db_info
        end_time = time.time()

        collected_test_info = current_test_sensor_data.get(test_id)
        if not collected_test_info:
            socketio.emit('test_stopped_ack', {
                'test_id': test_id,
                'success': False,
                'message': 'No data found for test ID (possibly already stopped).'
            })
            return

        collected_data = collected_test_info['data']
        sensor_data_json = json.dumps(collected_data)

        # Step 3: Update DB with end time and sensor data
        c.execute("UPDATE walk_tests SET end_time = ?, sensor_data_json = ? WHERE test_id = ?",
                  (end_time, sensor_data_json, test_id))
        conn.commit()
        conn.close()

        # Step 4: Save to Excel
        file_path = initialize_excel(patient_id)
        if file_path:
            wb = load_workbook(file_path)
            sheet_name = walk_type.upper().replace(' ', '_')

            if sheet_name not in wb.sheetnames:
                ws = wb.create_sheet(sheet_name)
                ws.append(["Timestamp"] + [f"S{i+1}" for i in range(12)])
            else:
                ws = wb[sheet_name]

            for i, row_data in enumerate(collected_data):
                if isinstance(row_data, list) and all(isinstance(x, (int, float)) for x in row_data):
                    timestamp = len(ws['A'])
                    ws.append([timestamp] + row_data)

            wb.save(file_path)
        else:
            socketio.emit('test_stopped_ack', {
                'test_id': test_id,
                'success': False,
                'message': f'Excel file not found or could not be created for patient {patient_id}'
            })
            return

        # Step 5: Clear in-memory buffer
        if test_id in current_test_sensor_data:
            del current_test_sensor_data[test_id]

        socketio.emit('test_stopped_ack', {
            'test_id': test_id,
            'success': True,
            'message': 'Test stopped successfully and data saved.'
        })

    except Exception as e:
        print(f"ERROR while stopping test {test_id}: {e}")
        socketio.emit('test_stopped_ack', {
            'test_id': test_id,
            'success': False,
            'message': f'Server error: {e}'
        })
    finally:
        if conn:
            conn.close()


@app.route('/save_chart_image', methods=['POST'])
@cross_origin() # Explicitly enable CORS for this route if global CORS(app) isn't enough
def save_chart_image():
    try:
        data = request.json
        doctor_username = data.get('doctor_username')
        patient_id = data.get('patient_id')
        walk_type = data.get('walk_type')
        foot_type = data.get('foot_type')
        image_data_b64 = data.get('image_data') # This will be the "data:image/png;base64,..." string

        if not all([doctor_username, patient_id, walk_type, foot_type, image_data_b64]):
            logging.error("Missing data for saving chart image.")
            return jsonify({'message': 'Missing data for chart image saving.'}), 400

        # Validate doctor and patient (similar to get_analyses_for_patient)
        conn = sqlite3.connect('spinalysis.db')
        c = conn.cursor()
        c.execute("SELECT first_name, last_name FROM patients WHERE patient_id = ? AND doctor_username = ?", (patient_id, doctor_username))
        patient_info = c.fetchone()
        conn.close()

        if not patient_info:
            logging.warning(f"Unauthorized or invalid patient/doctor for image save: patient_id={patient_id}, doctor={doctor_username}")
            return jsonify({'message': 'Unauthorized or invalid patient/doctor.'}), 403

        # Extract base64 part (remove "data:image/png;base64," prefix)
        header, encoded_data = image_data_b64.split(',', 1)
        image_bytes = base64.b64decode(encoded_data)

        # Determine save path for charts
        patient_first_name, patient_last_name = patient_info[0], patient_info[1]
        patient_folder = create_patient_folder(patient_first_name, patient_last_name) # Reuse your function

        charts_subfolder = os.path.join(patient_folder, 'charts')
        os.makedirs(charts_subfolder, exist_ok=True) # Create 'charts' subfolder if it doesn't exist

        # Define filename (e.g., Heel_Walk_left_chart.png)
        file_name = f"{walk_type}_{foot_type}_chart.png"
        file_path = os.path.join(charts_subfolder, file_name)

        # Save the image
        with open(file_path, 'wb') as f:
            f.write(image_bytes)

        logging.info(f"Chart saved successfully: {file_path}")
        saved_filename = os.path.basename(file_path) # Get just the filename (e.g., 'Heel_Walk_left_chart.png')
        return jsonify({'message': 'Chart image saved', 'filename': saved_filename}), 200

    except Exception as e:
        logging.error(f"Error saving chart image: {e}", exc_info=True)
        return jsonify({'message': f'Failed to save chart image: {str(e)}'}), 500



@app.route('/get_analyses_for_patient/<int:patient_id>', methods=['POST']) # Changed to POST to receive doctor_username in body
def get_analyses_for_patient(patient_id):
    """
    Fetches stored sensor data for ALL completed tests of a given patient
    and performs analysis based on sensor groups (Forefoot, Arch, Heel).
    Returns analysis data for all available walk types for the patient.
    Requires doctor_username in the request body for authorization.
    """
    data = request.json # Get data from request body to access doctor_username
    doctor_username = data.get('doctor_username')

    if not doctor_username:
        return jsonify({'message': 'Unauthorized: Doctor username missing in request'}), 401

    conn = sqlite3.connect('spinalysis.db')
    c = conn.cursor()
    
    # First, verify that the patient belongs to the current doctor
    c.execute("SELECT 1 FROM patients WHERE patient_id = ? AND doctor_username = ?", (patient_id, doctor_username))
    if c.fetchone() is None:
        conn.close()
        logging.warning(f"Unauthorized access: Patient {patient_id} does not belong to doctor {doctor_username}.")
        return jsonify({'message': 'Unauthorized: Patient does not belong to this doctor'}), 403

    # Retrieve all walk_tests for the given patient_id
    c.execute("SELECT test_id, walk_type, sensor_data_json FROM walk_tests WHERE patient_id = ?", (patient_id,))
    all_tests = c.fetchall()
    conn.close()

    if not all_tests:
        return jsonify({'message': 'No test records found for this patient.', 'filtered_samples': 0}), 200

    patient_analyses = {}
    total_filtered_samples_across_all_tests = 0

    for test_id, walk_type, sensor_data_json in all_tests:
        all_sensor_data = json.loads(sensor_data_json) if sensor_data_json else []

        if not all_sensor_data:
            logging.warning(f"No sensor data found for test_id {test_id} ({walk_type}). Skipping analysis for this test.")
            continue

        # Filtration Logic: Remove the first and last `samples_to_ignore` data points.
        samples_to_ignore = 20
        if len(all_sensor_data) > 2 * samples_to_ignore:
            filtered_data = all_sensor_data[samples_to_ignore:-samples_to_ignore]
        else:
            filtered_data = all_sensor_data

        if not filtered_data:
            logging.info(f"Test {test_id} ({walk_type}) has no valid samples after filtration. Skipping analysis for this test.")
            continue
            
        total_filtered_samples_across_all_tests += len(filtered_data)


        left_g1_sum, left_g2_sum, left_g3_sum = 0, 0, 0
        left_g1_count, left_g2_count, left_g3_count = 0, 0, 0

        right_g1_sum, right_g2_sum, right_g3_sum = 0, 0, 0
        right_g1_count, right_g2_count, right_g3_count = 0, 0, 0


        for sample_row in filtered_data:
            if len(sample_row) == 12:
                # Left Foot
                left_g1_sum += sum(sample_row[0:3]) # S1, S2, S3
                left_g1_count += 3

                left_g2_sum += sample_row[3] # S4
                left_g2_count += 1

                left_g3_sum += sum(sample_row[4:6]) # S5, S6
                left_g3_count += 2

                # Right Foot
                right_g1_sum += sum(sample_row[6:9]) # S7, S8, S9
                right_g1_count += 3

                right_g2_sum += sample_row[9] # S10
                right_g2_count += 1

                right_g3_sum += sum(sample_row[10:12]) # S11, S12
                right_g3_count += 2
            else:
                logging.warning(f"Skipping malformed sensor data sample (expected 12, got {len(sample_row)}) for test {test_id}: {sample_row}")
                continue

        avg_left_g1 = (left_g1_sum / left_g1_count) if left_g1_count > 0 else 0
        avg_left_g2 = (left_g2_sum / left_g2_count) if left_g2_count > 0 else 0
        avg_left_g3 = (left_g3_sum / left_g3_count) if left_g3_count > 0 else 0

        avg_right_g1 = (right_g1_sum / right_g1_count) if right_g1_count > 0 else 0
        avg_right_g2 = (right_g2_sum / right_g2_count) if right_g2_count > 0 else 0
        avg_right_g3 = (right_g3_sum / right_g3_count) if right_g3_count > 0 else 0
        
        patient_analyses[f"{walk_type}_analysis"] = {
            "left_foot_group_averages": [avg_left_g1, avg_left_g2, avg_left_g3],
            "right_foot_group_averages": [avg_right_g1, avg_right_g2, avg_right_g3]
        }
    
    patient_analyses['filtered_samples'] = total_filtered_samples_across_all_tests

    if not any(key.endswith('_analysis') for key in patient_analyses):
        return jsonify({'message': 'No valid analysis data could be generated for any walk type for this patient.', 'filtered_samples': 0}), 200
    try:
        excel_file_path = initialize_excel(patient_id)
        if excel_file_path:
            wb = load_workbook(excel_file_path)
            
            summary_sheet_name = "Analysis Summary"
            if summary_sheet_name in wb.sheetnames:
                ws = wb[summary_sheet_name]
                # Clear existing content if sheet exists and is not empty
                # Delete rows from 2nd row to max_row to clear data but keep headers
                for row in range(ws.max_row, 1, -1):
                    ws.delete_rows(row)
            else:
                ws = wb.create_sheet(summary_sheet_name)
            
            # Write headers for the analysis summary sheet
            analysis_headers = ["Walk Type", "Foot", "Forefoot Average", "Arch Average", "Heel Average"]
            ws.append(analysis_headers)

            # Populate the sheet with analysis data
            for walk_type_key in patient_analyses:
                if walk_type_key.endswith('_analysis'):
                    walk_type_name = walk_type_key.replace('_analysis', '').replace('_', ' ') # Make it human readable
                    analysis_data = patient_analyses[walk_type_key]

                    left_averages = analysis_data.get('left_foot_group_averages', [0, 0, 0])
                    right_averages = analysis_data.get('right_foot_group_averages', [0, 0, 0])

                    # Append Left Foot data
                    ws.append([walk_type_name, "Left", *left_averages])
                    # Append Right Foot data
                    ws.append([walk_type_name, "Right", *right_averages])
            
            # Add total filtered samples
            ws.append([]) # Blank row for separation
            ws.append(["Total Filtered Samples:", patient_analyses.get('filtered_samples', 0)])

            wb.save(excel_file_path)
            logging.info(f"Analysis summary saved to Excel for patient {patient_id} at {excel_file_path}")
        else:
            logging.warning(f"Could not save analysis summary to Excel for patient {patient_id}: Excel file path not found.")
    except Exception as e:
        logging.error(f"Error saving analysis summary to Excel for patient {patient_id}: {e}", exc_info=True)
    
    return jsonify(patient_analyses), 200


@app.route('/generate_graph_report', methods=['POST'])
@cross_origin() # Ensure CORS is enabled for this route
def generate_graph_report():
    try:
        data = request.json
        doctor_username = data.get('doctor_username')
        patient_id = data.get('patient_id')
        chart_image_filenames = data.get('chart_image_filenames') # This is the dictionary from frontend

        if not all([doctor_username, patient_id, chart_image_filenames]):
            logging.error("Missing data for graph report generation.")
            return jsonify({'message': 'Missing data for graph report generation.'}), 400

        # Validate doctor and patient (similar to other routes)
        conn = sqlite3.connect('spinalysis.db')
        c = conn.cursor()
        c.execute("SELECT first_name, last_name FROM patients WHERE patient_id = ? AND doctor_username = ?", (patient_id, doctor_username))
        patient_info = c.fetchone()
        conn.close()

        if not patient_info:
            logging.warning(f"Unauthorized or invalid patient/doctor for graph report: patient_id={patient_id}, doctor={doctor_username}")
            return jsonify({'message': 'Unauthorized or invalid patient/doctor.'}), 403

        # Extract patient details for the report
        patient_first_name, patient_last_name = patient_info

        # Reconstruct patient folder path
        patient_folder = create_patient_folder(patient_first_name, patient_last_name)
        charts_subfolder = os.path.join(patient_folder, 'charts')

        if not os.path.exists(charts_subfolder):
            logging.error(f"Charts folder not found for patient {patient_id}: {charts_subfolder}")
            return jsonify({'message': 'Chart images not found on server.'}), 404

        # --- PDF Generation Logic ---
        pdf = FPDF()
        pdf.set_auto_page_break(auto=True, margin=15) # Auto page break with 15mm margin
        pdf.add_page()
        pdf.set_font("Arial", "B", 16)
        pdf.cell(0, 10, "Patient Gait Analysis - Graph Report", 0, 1, "C")
        pdf.ln(10) # Line break

        # Patient Information Section
        pdf.set_font("Arial", "B", 12)
        pdf.cell(0, 7, "Patient Information:", 0, 1)
        pdf.set_font("Arial", "", 10)
        pdf.cell(0, 6, f"Name: {patient_first_name} {patient_last_name or ''}", 0, 1)
        pdf.cell(0, 6, f"Patient ID: {patient_id}", 0, 1)
        
        pdf.ln(10)

        # Doctor Information
        pdf.set_font("Arial", "B", 12)
        pdf.cell(0, 7, "Doctor Information:", 0, 1)
        pdf.set_font("Arial", "", 10)
        pdf.cell(0, 6, f"Username: {doctor_username}", 0, 1)
        pdf.ln(10)

        # Iterate through chart image filenames and add them to PDF
        # Order of walk types for consistent report
        walk_type_order = ['Heel_Walk', 'Normal_Walk', 'Forefeet_Walk']
        foot_type_order = ['left', 'right']

        # Determine image width and height for layout
        # Assuming charts are roughly square or slightly wider than tall
        # Adjust these values based on your actual chart dimensions and desired PDF layout
        img_width = 90 # mm
        img_height = 60 # mm
        margin_x = (pdf.w - (2 * img_width)) / 3 # Calculate margin to center two images
        margin_y = 10 # Vertical space between rows of images

        for walk_type in walk_type_order:
            # Check if any charts exist for this walk type
            has_charts_for_walk_type = False
            for foot_type in foot_type_order:
                key = f"{walk_type}_{foot_type}"
                if key in chart_image_filenames:
                    has_charts_for_walk_type = True
                    break
            
            if not has_charts_for_walk_type:
                continue # Skip if no charts for this walk type

            pdf.add_page() # Start a new page for each walk type
            pdf.set_font("Arial", "B", 14)
            pdf.cell(0, 10, f"{walk_type.replace('_', ' ')} Analysis", 0, 1, "C")
            pdf.ln(5)

            # Keep track of current X position for side-by-side images
            current_x = margin_x

            for foot_type in foot_type_order:
                key = f"{walk_type}_{foot_type}"
                filename = chart_image_filenames.get(key)
                
                if filename:
                    image_path = os.path.join(charts_subfolder, filename)
                    if os.path.exists(image_path):
                        # Add image label
                        pdf.set_font("Arial", "B", 10)
                        pdf.set_x(current_x) # Position label above image
                        pdf.cell(img_width, 7, f"{foot_type.capitalize()} Foot", 0, 0, "C")
                        pdf.ln(7) # Move down for image

                        # Add image
                        pdf.set_x(current_x) # Position image
                        pdf.image(image_path, x=current_x, w=img_width, h=img_height)
                        
                        current_x += img_width + margin_x # Move X for next image
                    else:
                        logging.warning(f"Chart image not found: {image_path}")
                        # Optionally add a placeholder text if image is missing
                        pdf.set_font("Arial", "I", 10)
                        pdf.set_x(current_x)
                        pdf.cell(img_width, img_height, f"Image missing: {filename}", 1, 0, "C")
                        current_x += img_width + margin_x
                else:
                    # If no filename for this foot type, still advance X to maintain layout
                    current_x += img_width + margin_x

            pdf.ln(img_height + margin_y) # Move down for next row/section if needed on same page
            # If you want each chart on its own page, remove the current_x logic and just add_page() before each image.

        # Output PDF as bytes
        pdf_output = pdf.output(dest='S') # Will return bytes or bytearray directly, no need to encode

        # Return the PDF as a file download
        response = app.make_response(pdf_output)
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = f'attachment; filename=Graph_Report_{patient_first_name}_{patient_last_name or ""}_{time.strftime("%Y%m%d-%H%M%S")}.pdf'
        return response

    except Exception as e:
        logging.error(f"Error generating graph report: {e}", exc_info=True)
        return jsonify({'message': f'Failed to generate graph report: {str(e)}'}), 500



if __name__ == '__main__':
    # This ensures the database is initialized before the app starts
    init_db()

    # --- Initialize Serial Port and Start Sensor Reading Thread ---
    if sensor_reading_thread is None:
        print("INFO: Initializing sensor reading thread...")
        sensor_reading_thread = threading.Thread(target=read_from_serial_port, daemon=True)
        sensor_reading_thread.start()
        print("INFO: Sensor reading thread started.")

    # ✅ Start server using eventlet (thanks to monkey_patch already done)
    socketio.run(app, host='0.0.0.0', port=8000)
