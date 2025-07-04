import serial
from serial.tools import list_ports
import time
import os
from openpyxl import Workbook, load_workbook # Ensure load_workbook is imported here
from openpyxl.styles import Font, Alignment
from datetime import datetime

# --- Configuration ---
BAUD_RATE = 115200 # Must match your ESP32's baud rate
EXCEL_OUTPUT_DIR = 'dummy_sensor_data' # Folder to store Excel files
EXCEL_FILE_PREFIX = 'dummy_test_data' # Prefix for Excel filenames
DURATION_SECONDS = 10 # How long the dummy test should run (in seconds)

# Ensure the output directory exists
os.makedirs(EXCEL_OUTPUT_DIR, exist_ok=True)

# --- Helper Function for Serial Port Auto-Detection (Copied from your app.py) ---
def auto_detect_esp_port():
    """
    Attempts to auto-detect the serial port of an ESP32.
    It looks for common Vendor IDs (VIDs) and Product IDs (PIDs) associated with ESP32s.
    """
    CP210X_VID = 0x10C4 # Silicon Labs
    CP210X_PID = 0xEA60 # CP2102/CP2104
    CH340_VID = 0x1A86 # QinHeng Electronics
    CH340_PID = 0x7523 # CH340G / CH341A
    ESP32_C3_VID = 0x303A
    ESP32_C3_PID = 0x1001

    found_port = None
    print("INFO: Attempting to auto-detect ESP32 serial port...")

    for port in list_ports.comports():
        print(f"DEBUG: Found port: {port.device}, Desc: {port.description}, HWID: {port.hwid}, VID: {port.vid}, PID: {port.pid}")

        if (port.vid == CP210X_VID and port.pid == CP210X_PID) or \
           (port.vid == CH340_VID and port.pid == CH340_PID) or \
           (port.vid == ESP32_C3_VID and port.pid == ESP32_C3_PID):
            print(f"INFO: Auto-detected ESP32 port (VID/PID match): {port.device}")
            found_port = port.device
            break

        if "CP210" in port.description or "USB-SERIAL CH340" in port.description or "ESP32-C3" in port.description:
            print(f"INFO: Auto-detected ESP32 port (Description match): {port.device}")
            found_port = port.device
            break

    if found_port:
        return found_port
    else:
        print("WARNING: No ESP32 serial port auto-detected based on common VIDs/PIDs or descriptions.")
        print("Please ensure your Master ESP32 is connected and its drivers are installed.")
        return None

# --- Function to write data to Excel ---
def append_to_excel(file_path, header_needed, timestamp, raw_sensor_values):
    """
    Appends a row of sensor data to an Excel file.
    """
    if header_needed or not os.path.exists(file_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Sensor Data"
        headers = ["Timestamp"] + [f"S{i}" for i in range(1, 13)]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
    else:
        # load_workbook must be available in this scope, which it is via global import
        wb = load_workbook(file_path)
        ws = wb.active

    data_row = [timestamp] + raw_sensor_values
    ws.append(data_row)
    wb.save(file_path)

# --- Main Dummy Test Function ---
def run_dummy_test():
    com_port = auto_detect_esp_port()
    if com_port is None:
        print("ERROR: No ESP32 serial port found. Exiting dummy test.")
        return

    ser = None
    try:
        ser = serial.Serial(com_port, BAUD_RATE, timeout=1)
        print(f"INFO: Serial port {com_port} opened successfully at {BAUD_RATE} baud.")
        time.sleep(2) # Give ESP time to reset

        # Send "start" command
        command_to_send = "start\n"
        ser.write(command_to_send.encode('utf-8'))
        print(f"INFO: Sent '{command_to_send.strip()}' command to ESP32.")

        excel_filename = os.path.join(EXCEL_OUTPUT_DIR, f"{EXCEL_FILE_PREFIX}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
        header_needed = True # Flag to add header only for the first save

        start_time = time.time()
        print(f"\n--- Starting Dummy Sensor Read for {DURATION_SECONDS} seconds ---")

        while (time.time() - start_time) < DURATION_SECONDS:
            if ser.in_waiting > 0:
                line = ser.readline().decode('utf-8').strip()
                if line:
                    print(f"RAW: '{line}'") # Display raw data in terminal

                    values_str = line.split(',')
                    if len(values_str) == 12:
                        try:
                            # Convert to float as per your ESP32's likely output (e.g., 4095.0)
                            raw_values = [float(v) for v in values_str]
                            print(f"PARSED: {raw_values}") # Display parsed data in terminal

                            current_timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S.%f')[:-3]
                            append_to_excel(excel_filename, header_needed, current_timestamp, raw_values)
                            header_needed = False # Header only needed for the first save

                        except ValueError as ve:
                            print(f"WARNING: Could not convert data to float: {ve} in line: '{line}'")
                        except Exception as e:
                            print(f"ERROR: General error during data processing: {e}")
                    else:
                        print(f"WARNING: Unexpected number of values ({len(values_str)}) in line: '{line}'")
            time.sleep(0.01) # Small delay to prevent busy-waiting

        print("\n--- Dummy Sensor Read Finished ---")

    except serial.SerialException as e:
        print(f"ERROR: Serial communication error: {e}")
    except Exception as e:
        print(f"CRITICAL ERROR: {e}")
    finally:
        if ser and ser.is_open:
            # Send "nope" command to stop data flow
            command_to_send = "nope\n"
            ser.write(command_to_send.encode('utf-8'))
            print(f"INFO: Sent '{command_to_send.strip()}' command to ESP32.")
            ser.close()
            print(f"INFO: Serial port {com_port} closed.")

if __name__ == '__main__':
    run_dummy_test()